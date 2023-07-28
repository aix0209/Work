#!/usr/bin/env python
# -*- coding: utf-8 -*-
#2018/01/22 排列課表以教師中心為排課對象
#2018/02/05 修正連兩堂排課功能，加入連3、連4堂
    #排列問題仍無法解決，到最後一科歷史時空堂空間太少，重排多次仍無法改善此問題
    #有可能解決辦法有:
    #1.將選星期從最多空堂改為隨機方式選取，但之前試過效果比較不好
    #2.排課時從有連兩堂空堂的排起，但仍無法解決問題
    #3.加入自動調課機制，但判斷及寫法比較困難
    #4.可嘗試從最少空堂的排此，但預期效果不好(2018/02/05已試過無法解決問題)
    #5.觀察到的問題是在排列時優先排課的課程大多從星期一開始排起，原因不明
#2018/02/05 開始測試連4堂功能，發現之前的選擇日期的方法會有問題，若只有一天則無法隨機選擇
#2018/02/11 發現單一排課會有周一天重覆排課情形，嘗試修改隨機行為目前沒結果
    #重覆排課情形可能來自於，選擇最多空白數那天排課後，刪掉那天的index跟重新取有最大值的index不一致
    #第二個原因來自於在做優先2空堂排課時沒有將同堂名排除
#2018/02/13 解決重覆排課問題，問題出現在堂名比較時沒有使用 字串 in而用==會不對
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment
import random
from operator import itemgetter, attrgetter
import operator
import math
import os
import time
def C(a,b):#排列組合
    a=int(a)
    b=int(b)
    return ByNum(a)/(ByNum(b)*ByNum(a-b))

def ByNum(Num):#階層
    ByTotal=1
    if Num!=0:
        for LoopNumA in range(1,Num+1):
            ByTotal*=LoopNumA
    else:
        ByTotal=1
    return ByTotal

def ClassForm():#建立課表
    ClassTemp=[[1 for col in range(6)]for row in range(8)]
    return ClassTemp


class ExcelAction:
    def __init__(self,FileName):#啟始變數
        self.SeriaWord='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.FileName=''
        self.FileName=FileName
        self.row=0
        self.col=0
        self.Dim=0
        ColNames=''
        self.Data=load_workbook(self.FileName,data_only = True)
        self.SheetNames=self.Data.get_sheet_names()
        self.OneCount=0

    def Loading(self,SheetNO):#載入指定Excel,row值,col值
        self.Sheet=self.Data.get_sheet_by_name(self.Data.get_sheet_names()[SheetNO])
        self.row=self.Sheet.max_row#取row值
        self.col=self.Sheet.max_column#取col值
        return self.SheetNames[SheetNO],self.Sheet,self.row,self.col


class RangeTime:
    def __init__(self,TeacherData,ClassTemp,ClassData,LessonData,TeacherDataOrigin):#啟始變數
        self.a=0
        self.TeacherData=TeacherData
        self.TeacherIndex=0
        self.ClassTemp=ClassTemp
        self.ClassData=ClassData
        self.LessonData=LessonData
        self.TeacherDataOrigin=TeacherDataOrigin
        self.LessonInfo=[]


    def reNewClass(self,ClassIndex):
        for LoopNumA in range(1,6,1):
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]=7
        for LoopNumA1 in range(1,6,1):
            for LoopNumB in range(1,8,1):
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA1]=1


    def reSetClass(self,ClassIndex):
        for LoopNumA1 in range(1,6,1):
            for LoopNumB in range(1,8,1):
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA1]==0:
                  self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA1]=1

#############################################################################################################
    def RangeLesson1(self,LessonName,LessonTime,ClassNO,ClassIndex):
        JugeNum=0
        CountET=0
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        while LessonTime>=6:
            for LoopNumA in range(1,6,1):
                CountET=0
                JugeNum=1
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=2:#如果當天還有至少二堂的空堂，就做判定
                    for LoopNumB in (1,2,3,5,6):#判定是否有同樣科目
                        if LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] )or \
                        LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] ):
                           JugeNum=0

                    if JugeNum!=0:#若無同樣科目
                        for LoopNumB in (1,2,3,5,6):#紀錄空堂位置
                            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and\
                            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1 :
                               CountET+=1#紀錄空堂次數，以2堂為一次
                        if CountET>=1:
                            RGDayRecordIndex.append(LoopNumA)#紀錄星期
                            RGDayRecordTimes.append(CountET)#紀錄空堂數
            LessonRangeTimeRecord=[]
            MaxIndex=RGDayRecordTimes.index(max(RGDayRecordTimes))
            LessonRangeDay=RGDayRecordIndex[MaxIndex]#找空堂數最多的星期
            del RGDayRecordIndex[MaxIndex]#刪掉該日
            del RGDayRecordTimes[MaxIndex]#刪掉該日的空堂紀錄
            #LessonRangeDay=random.choice(RGDayRecordIndex)#隨機選取
            #RGDayRecordIndex.remove(LessonRangeDay)#隨機選取移除
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=2
            for LoopNumA in (1,2,3,5,6):#排二堂課到當天,只到第六節
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA][LessonRangeDay]==1 and\
                 self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA+1][LessonRangeDay]==1:
                    LessonRangeTimeRecord.append(LoopNumA)
            random.seed(time.time())#隨機種子
            TimeIndex=random.randint(0,len(LessonRangeTimeRecord)-1)
            #LessonRangeTime=random.choice(LessonRangeTimeRecord)
            LessonRangeTime=LessonRangeTimeRecord[TimeIndex]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            if LessonName!='班級活動':
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('節數')]-=2
            LessonTime-=2

        #只安排一堂課，選空堂最多的星期，及有連兩堂空堂的天數為優先安排
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        for LoopNumA in range(1,6,1):#先從有連兩堂空堂排起
            JugeNum=1
            CountET=0

            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=2:#如果當天還有至少二堂的空堂，就做判定
                for LoopNumB in (1,2,3,5,6):#判定是否有同樣科目
                    if LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] )or \
                    LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] ):
                       JugeNum=0
                if JugeNum!=0:#若無同樣科目
                    for LoopNumB in (1,2,3,5,6):#紀錄空堂位置
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and\
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1 :
                           CountET+=1#紀錄空堂次數，以2堂為一次
                    if CountET>=1:
                        RGDayRecordIndex.append(LoopNumA)#紀錄星期
                        RGDayRecordTimes.append(CountET)#紀錄空堂數

        if RGDayRecordIndex!=[]:


            for LoopNumLesson in range(LessonTime):
                LessonRangeTimeRecord=[]
                #LessonRangeDay=random.choice(RGDayRecordIndex)#隨機選取
                #RGDayRecordIndex.remove(LessonRangeDay)#隨機選取移除
                MaxIndex=RGDayRecordTimes.index(max(RGDayRecordTimes))
                LessonRangeDay=RGDayRecordIndex[MaxIndex]#找空堂數最多的星期
                del RGDayRecordIndex[MaxIndex]#刪掉該日
                del RGDayRecordTimes[MaxIndex]#刪掉該日的空堂紀錄
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=1#扣當天一空堂

                for LoopNumA in (1,2,3,5,6):##排一堂課到當天
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA][LessonRangeDay]==1 and \
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA+1][LessonRangeDay]==1 :
                        LessonRangeTimeRecord.append(LoopNumA)
                random.seed()#隨機種子
                TimeIndex=random.randint(0,len(LessonRangeTimeRecord)-1)
                print(ClassNO, TimeIndex)
                #LessonRangeTime=random.choice(LessonRangeTimeRecord)
                LessonRangeTime=LessonRangeTimeRecord[TimeIndex]
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
                if LessonName!='班級活動':
                    self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                    self.TeacherData[self.TeacherIndex][TeacherData[0].index('節數')]-=1
                LessonTime-=1
                if RGDayRecordIndex==[]:
                    break
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        for LoopNumA in range(1,6,1):#再排剩下堂數
            JugeNum=1
            CountET=0
            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=1:#如果當天還有至少一堂的空堂，就做判定
                for LoopNumB in range(1,8,1):#判定是否有同樣科目
                    if LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] ):
                        JugeNum=0
                if JugeNum!=0:#若無同樣科目
                    for LoopNumB in range(1,8,1):#紀錄空堂位置
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                           CountET+=1#紀錄空堂次數，以1堂為一次
                    if CountET>=1:
                        RGDayRecordIndex.append(LoopNumA)#紀錄星期
                        RGDayRecordTimes.append(CountET)#紀錄空堂數

        for LoopNumLesson in range(LessonTime):
            LessonRangeTimeRecord=[]

            MaxIndex=RGDayRecordTimes.index(max(RGDayRecordTimes))
            LessonRangeDay=RGDayRecordIndex[MaxIndex]#找空堂數最多的星期
            del RGDayRecordIndex[MaxIndex]#刪掉該日
            del RGDayRecordTimes[MaxIndex]#刪掉該日的空堂紀錄
            #LessonRangeDay=random.choice(RGDayRecordIndex)#隨機選取
            #RGDayRecordIndex.remove(LessonRangeDay)#隨機選取移除
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=1#扣當天一空堂
            for LoopNumA in range(1,8,1):#排一堂課到當天
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA][LessonRangeDay]==1:
                    LessonRangeTimeRecord.append(LoopNumA)

            random.seed(time.time())#隨機種子
            TimeIndex=random.randint(0,len(LessonRangeTimeRecord)-1)
            #LessonRangeTime=random.choice(LessonRangeTimeRecord)
            LessonRangeTime=LessonRangeTimeRecord[TimeIndex]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            if LessonName!='班級活動':
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('節數')]-=1
            LessonTime-=1
        self.ClassData[ClassIndex][self.ClassData[0].index('科目教師')].append(LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')])
#########################################################################################################################
    def RangeLesson2(self,LessonName,LessonTime,ClassNO,ClassIndex):#連2堂
        JugeNum=0
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        for LoopNumA in range(1,6,1):
            JugeNum=1
            CountET=0
            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=2:#如果當天還有至少二堂的空堂，且當日無相同科目，就加入排列
                for LoopNumB in range(1,7,1):
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1:
                       CountET+=1
                    if LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA] ) and \
                    LessonName in str( self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA] ):
                       JugeNum=0
                       break
                if JugeNum==1 and CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)

        for LoopNumLesson in range(0,LessonTime,2):
            LessonRangeTimeRecord=[]
            LessonRangeDay=RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))]#找空堂數最多的星期
            RGDayRecordIndex.remove(RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))])
            RGDayRecordTimes.remove(max(RGDayRecordTimes))
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=2

            for LoopNumA in range(1,7,1):#排二堂課到當天,只到第六節
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA][LessonRangeDay]==1 and\
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA+1][LessonRangeDay]==1:
                    LessonRangeTimeRecord.append(LoopNumA)
            LessonRangeTime=random.choice(LessonRangeTimeRecord)
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            if LessonName!='班級活動':
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=ClassNO
#####################################################
######################################################

#########################################################################################################################
    def RangeLesson3(self,LessonName,LessonTime,ClassNO,ClassIndex):#連3堂
        JugeNum=0
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        for LoopNumA in range(1,6,1):
            JugeNum=1
            CountET=0
            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=3:#如果當天還有至少二堂的空堂，且當日無相同科目，就加入排列
                for LoopNumB in (1,2,5):
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1 and \
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+2][LoopNumA]==1:
                       CountET+=1
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==LessonName:
                       JugeNum=0
                       break
                if JugeNum==1 and CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                #for LoopNumB in range(1,7,1):
                   # if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                    #self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1:
                    #    RGDayRecordIndex.append(LoopNumA)
                    #    RGDayRecordTimes.append(CountET)
                    #    break
        for LoopNumLesson in range(0,LessonTime,3):
            LessonRangeTimeRecord=[]
            LessonRangeDay=RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))]#找空堂數最多的星期
            RGDayRecordIndex.remove(RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))])
            RGDayRecordTimes.remove(max(RGDayRecordTimes))
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=3
            for LoopNumA in (1,2,5):#排三堂課到當天,只到第六節
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA][LessonRangeDay]==1 and\
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA+1][LessonRangeDay]==1 and \
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumA+2][LessonRangeDay]==1:
                    LessonRangeTimeRecord.append(LoopNumA)
            LessonRangeTime=random.choice(LessonRangeTimeRecord)
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+2][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            if LessonName!='班級活動':
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+2][LessonRangeDay]=ClassNO
#####################################################
######################################################

    def RangeLesson4(self,LessonName,LessonTime,ClassNO,ClassIndex):#連4堂
        JugeNum=0
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        for LoopNumA in range(1,6,1):
            JugeNum=1
            CountET=0
            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LoopNumA]>=4:#如果當天還有至少四堂的空堂，且當日無相同科目，就加入排列
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1][LoopNumA]==1 and \
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+1][LoopNumA]==1 and \
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+2][LoopNumA]==1 and \
                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+3][LoopNumA]==1:
                   CountET+=1
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1][LoopNumA]==LessonName:
                   JugeNum=0
                   break

                if JugeNum==1 and CountET>0:
                    RGDayRecordIndex.append(LoopNumA)
                    RGDayRecordTimes.append(CountET)

        for LoopNumLesson in range(0,LessonTime,4):
            LessonRangeTimeRecord=[]
            #LessonRangeDay=random.choice(RGDayRecordIndex)#隨機選取
            #RGDayRecordIndex.remove(LessonRangeDay)#隨機選取移除
            LessonRangeDay=RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))]#找空堂數最多的星期
            RGDayRecordIndex.remove(RGDayRecordIndex[RGDayRecordTimes.index(max(RGDayRecordTimes))])
            RGDayRecordTimes.remove(max(RGDayRecordTimes))
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][0][LessonRangeDay]-=4

            #排四堂課到當天,一到四節
            LessonRangeTime=1
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+2][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LessonRangeTime+3][LessonRangeDay]=LessonName+','+self.TeacherData[self.TeacherIndex][TeacherData[0].index('姓名')]
            if LessonName!='班級活動':
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+1][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+2][LessonRangeDay]=ClassNO
                self.TeacherData[self.TeacherIndex][TeacherData[0].index('課表')][LessonRangeTime+3][LessonRangeDay]=ClassNO
#####################################################
######################################################

    def LessonConditionSet(self,LessonIndex,ClassIndex):#課表條件

########################上下午排課

        setTime=self.LessonData[LessonIndex][self.LessonData[0].index('上下午')]#判斷上下午是否可排課
        ClassSet=[[0 for col in range(6)]for row in range(8)]
        if setTime=='上午':
            ClassSet=[[0 for col in range(6)]for row in range(8)]
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,5,1):#上午第一節到第四節設1
                    ClassSet[LoopNumB][LoopNumA]=1
        elif setTime=='下午':
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(5,8,1):#下午第五節到第七節設1
                    ClassSet[LoopNumB][LoopNumA]=1
        else:
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#下午第五節到第七節設1
                    ClassSet[LoopNumB][LoopNumA]=1
        for LoopNumA in range(1,6,1):#星期
            for LoopNumB in range(1,8,1):#邏輯判定
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0 or self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and ClassSet[LoopNumB][LoopNumA]==1 :
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=0
########################星期排課
        setTime=self.LessonData[LessonIndex][self.LessonData[0].index('限制星期')]#判斷星期是否可排課
        setTime=str(setTime)
        ClassSet=[[0 for col in range(6)]for row in range(8)]
        if setTime!='12345':
            for LoopNumA in (setTime):
                for LoopNumB in range(1,8,1):#星期整天設1
                    ClassSet[LoopNumB][int(LoopNumA)]=1
        else:
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#下午第五節到第七節設1
                    ClassSet[LoopNumB][LoopNumA]=1
        for LoopNumA in range(1,6,1):#星期
            for LoopNumB in range(1,8,1):#邏輯判定
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0 or self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and ClassSet[LoopNumB][LoopNumA]==1:
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=0
######################節數排課
        setTime=self.LessonData[LessonIndex][self.LessonData[0].index('限制節')]#判斷星期是否可排課
        setTime=str(setTime)
        ClassSet=[[0 for col in range(6)]for row in range(8)]
        if setTime!='1234567':
            for LoopNumA in (setTime):
                for LoopNumB in range(1,6,1):#該節設1
                    ClassSet[int(LoopNumA)][LoopNumB]=1
        else:
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#下午第五節到第七節設1
                    ClassSet[LoopNumB][LoopNumA]=1

        for LoopNumA in range(1,6,1):#星期
            for LoopNumB in range(1,8,1):#邏輯判定
                 if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0 or self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                    ClassSet[LoopNumB][LoopNumA]==1:
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=0
######################特別排課
        for SC in range(0,10,1):
            setTime=str(self.LessonData[LessonIndex][self.LessonData[0].index('條件'+str(SC+1))])

            SCL=setTime.split('0')
            if SCL[0]!=0 or SCL[0]!=[] or SCL[0]!='None':
                if SCL[0]=='可':
                    ClassSet=[[0 for col in range(6)]for row in range(8)]
                    for SCLDay in range(1,len(SCL),2):
                        for SCLTime in SCL[SCLDay+1]:
                            ClassSet[int(SCLTime)][int(SCLDay)]=1

                elif SCL[0]=='不':
                    ClassSet=[[1 for col in range(6)]for row in range(8)]
                    for SCLDay in range(1,len(SCL),2):
                        for SCLTime in SCL[SCLDay+1]:
                            ClassSet[int(SCLTime)][int(SCL[int(SCLDay)])]=0
                for LoopNumA in range(1,6,1):#星期
                    for LoopNumB in range(1,8,1):#邏輯判定
                         if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0 or \
                         self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                            if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                            ClassSet[LoopNumB][LoopNumA]==1:
                                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                            else:
                                self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=0


    def reRangeLesson(self):#依權數排列課程
        self.LessonData[0].append('權數')
        for LoopNumA in range(1,len(self.LessonData)):
            self.LessonData[LoopNumA].append('')
        reLesson=[]
        TempList=[]
        for LoopNumA in range(1,len(self.LessonData),1):
            power=1
            #連堂數權數
            if self.LessonData[LoopNumA][self.LessonData[0].index('連堂數')]==4:
                power*=C(1,1)
            elif self.LessonData[LoopNumA][self.LessonData[0].index('連堂數')]==3:
                power*=C(3,1)
            elif self.LessonData[LoopNumA][self.LessonData[0].index('連堂數')]==2:
                power*=C(6,1)
            else:
                power*=C(7,1)
            if self.LessonData[LoopNumA][self.LessonData[0].index('上下午')]=='上午':#上下午權數
               # power+=20
                power*=C(4,1)
            elif self.LessonData[LoopNumA][self.LessonData[0].index('上下午')]=='下午':
                #power+=15
                power*=C(3,1)
            else:
                power*=C(7,1)

            if self.LessonData[LoopNumA][self.LessonData[0].index('限制星期')]=='12345':#限制星期權數
                if self.LessonData[LoopNumA][self.LessonData[0].index('節數')]<=5:
                    power*=C(5,self.LessonData[LoopNumA][self.LessonData[0].index('節數')])
                else:
                    power*=C(5,5)
            else:
                if self.LessonData[LoopNumA][self.LessonData[0].index('節數')]<=5:
                    power*=C(int(math.log(self.LessonData[LoopNumA][self.LessonData[0].index('限制星期')])+1)\
                    ,self.LessonData[LoopNumA][self.LessonData[0].index('節數')])


            if self.LessonData[LoopNumA][self.LessonData[0].index('限制節')]=='1234567':#限制節
               # power+=35
               power*=C(7,1)
            else:
                power*=C(int(math.log(self.LessonData[LoopNumA][self.LessonData[0].index('限制節')])+1),1)

            for condiLoop in range(1,11,1):
                SCL=str(self.LessonData[LoopNumA][self.LessonData[0].index('條件'+str(condiLoop))]).split('0')
                if self.LessonData[LoopNumA][self.LessonData[0].index('節數')]<=5:
                    if self.LessonData[LoopNumA][self.LessonData[0].index('條件'+str(condiLoop))]!=0:
                        if SCL[0]=='可':
                           day=len(SCL)-1
                           day=int(day/2)
                           power*=C(5,day)*C(day,self.LessonData[LoopNumA][self.LessonData[0].index('節數')])
                        elif SCL[0]=='不':
                           day=len(SCL)-1
                           day=5-int(day/2)
                           power*=C(5,day)*C(day,self.LessonData[LoopNumA][self.LessonData[0].index('節數')])
                    else:
                        power*=C(5,self.LessonData[LoopNumA][self.LessonData[0].index('節數')])
                else:
                    power*=C(5,5)
            self.LessonData[LoopNumA][self.LessonData[0].index('權數')]=(power/35**5)**0.2
        PowerIndex=self.LessonData[0].index('權數')
        TempList=self.LessonData[0]
        del self.LessonData[0]
        self.LessonData=sorted(self.LessonData, key=itemgetter(PowerIndex))
        self.LessonData.insert(0,TempList)
        return self.LessonData

    def reRangeTeacherData(self):#依權數排列老師
        TempList=[]
        reTeacher=[]
        self.TeacherData[0].append('權數')

        for LoopNumA in range(1,len(self.TeacherData),1):
            self.TeacherData[LoopNumA].append('')

        for LoopNumA in range(1,len(self.TeacherData),1):
            power=0

            power+=35/(self.TeacherData[LoopNumA][self.TeacherData[0].index('節數')])
            power+=35/self.TeacherData[LoopNumA][self.TeacherData[0].index('連堂數')]#連堂數權數

            if self.TeacherData[LoopNumA][self.TeacherData[0].index('上下午')]=='上午':#上下午權數
                power+=20
            elif self.TeacherData[LoopNumA][self.TeacherData[0].index('上下午')]=='下午':
                power+=15
            else:
                power+=35

            if self.TeacherData[LoopNumA][self.TeacherData[0].index('限制星期')]=='12345':#限制星期權數
                power+=35
            else:
                power+=int(math.log(self.TeacherData[LoopNumA][self.TeacherData[0].index('限制星期')])+1)*35

            if self.TeacherData[LoopNumA][self.TeacherData[0].index('限制節')]=='1234567':#限制星期權數
                power+=35
            else:
                power+=int(math.log(self.TeacherData[LoopNumA][self.TeacherData[0].index('限制節')])+1)*5

            self.TeacherData[LoopNumA][self.TeacherData[0].index('權數')]=power/35
        TempList=self.TeacherData[0]
        PowerIndex=self.TeacherData[0].index('權數')
        del self.TeacherData[0]
        self.TeacherData=sorted(self.TeacherData, key=itemgetter(PowerIndex))
        self.TeacherData.insert(0,TempList)

        return self.TeacherData

    def buildTeacherClassForm(self):
        self.TeacherData[0].append('課表')
        for LoopNumA in range(1,len(self.TeacherData)):
            self.TeacherData[LoopNumA].append([[1 for col in range(6)]for row in range(8)])


    def buildClassClassForm(self):
        self.ClassData[0].append('科目教師')
        for LoopNumA in range(1,len(self.ClassData)):
            self.ClassData[LoopNumA].append([])
        self.ClassData[0].append('課表')
        for LoopNumA in range(1,len(self.ClassData)):
            self.ClassData[LoopNumA].append([[1 for col in range(6)]for row in range(8)])
        for LoopNumA in range(1,len(self.ClassData)):
            for LoopNumB in range(1,6,1):
                self.ClassData[LoopNumA][self.ClassData[0].index('課表')][0][LoopNumB]=7

    def getTeacher(self,LessonName,LessonTime,LessonYears,LessonClass):#取符合條件的老師
        TeacherIndex=[]
        TeacherTime=[]
        SingleIndex=0
        index=0
        TF=False
        for LoopNumA in range(1,len(self.TeacherData),1):
            if self.TeacherData[LoopNumA][self.TeacherData[0].index('年級')]==LessonYears :
                if self.TeacherData[LoopNumA][self.TeacherData[0].index('班級')]==LessonClass or\
                self.TeacherData[LoopNumA][self.TeacherData[0].index('班級')]==0:
                    if self.TeacherData[LoopNumA][self.TeacherData[0].index('科目')]==LessonName and \
                       self.TeacherData[LoopNumA][self.TeacherData[0].index('節數')]-LessonTime>=0:
                        TeacherIndex.append(LoopNumA)
                        TeacherTime.append(self.TeacherData[LoopNumA][self.TeacherData[0].index('節數')])
        if len(TeacherIndex)<=2:#加入跨年級教師
        #if TeacherIndex==[]:

            for LoopNumA in range(1,len(self.TeacherData),1):

                if self.TeacherData[LoopNumA][self.TeacherData[0].index('年級')]==0:
                    if self.TeacherData[LoopNumA][self.TeacherData[0].index('班級')]==LessonClass or\
                    self.TeacherData[LoopNumA][self.TeacherData[0].index('班級')]==0:

                        if self.TeacherData[LoopNumA][self.TeacherData[0].index('科目')]==LessonName and \
                           self.TeacherData[LoopNumA][self.TeacherData[0].index('節數')]-LessonTime>=0:
                            TeacherIndex.append(LoopNumA)
                            TeacherTime.append(self.TeacherData[LoopNumA][self.TeacherData[0].index('節數')])
        self.TeacherIndex=random.choice(TeacherIndex)

        return self.TeacherIndex

    def reTeacher(self,LessonTime):#回復教師資料
        self.TeacherData[self.TeacherIndex][self.TeacherData[0].index('節數')]+=LessonTime

    def reLessonAllTeacher(self,LessonName,LessonYears,ClassYears):#回復該科教師節數資料
        LessonTime=0
        for TeacherIndex in range(1,len(self.TeacherData)):
            if self.TeacherData[TeacherIndex][self.TeacherData[0].index('科目')]==LessonName:

                if self.TeacherData[TeacherIndex][self.TeacherData[0].index('年級')]==LessonYears or \
                    self.TeacherData[TeacherIndex][self.TeacherData[0].index('年級')]==0:

                    for LoopNumA in range(1,6,1):
                        for LoopNumB in range(1,8,1):
                            if str(ClassYears)+'年' in str(self.TeacherData[TeacherIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]):
                                self.TeacherData[TeacherIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1
                                self.TeacherData[TeacherIndex][self.TeacherData[0].index('節數')]+=1

    def reNewLessonAllTeacher(self):#回復該科教師節數資料
        for TeacherIndex in range(1,len(self.TeacherData)):
            for TeacherIndexOrigin in range(1,len(self.TeacherDataOrigin)):#找原始檔同樣科目的節數
                if self.TeacherDataOrigin[TeacherIndexOrigin][self.TeacherDataOrigin[0].index('科目')]==\
                self.TeacherData[TeacherIndex][self.TeacherData[0].index('科目')]:
                    self.TeacherData[TeacherIndex][self.TeacherData[0].index('節數')]=\
                    self.TeacherDataOrigin[TeacherIndexOrigin][self.TeacherDataOrigin[0].index('節數')]
                    break
            for LoopNumA in range(1,6,1):
                for LoopNumB in range(1,8,1):
                    self.TeacherData[TeacherIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1

    def reLessonAllClass(self,LessonName,LessonYears):#回復該科各班課表資料
        for ClassIndx in range(1,len(self.ClassData)):
            if len( self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')] )!=0:
                if LessonName in self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')][len( self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')] )-1]:
                    del self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')][len(self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')] )-1]
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#上午第一節到第四節設1
                    if LessonName in str(self.ClassData[ClassIndx][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]) and\
                    self.ClassData[ClassIndx][self.ClassData[0].index('年級')]==LessonYears:
                        self.ClassData[ClassIndx][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                        self.ClassData[ClassIndx][self.ClassData[0].index('課表')][0][LoopNumA]+=1

    def reNewLessonAllClass(self):#回復該科各班課表資料
        for ClassIndx in range(1,len(self.ClassData)):
            self.ClassData[ClassIndx][self.ClassData[0].index('科目教師')]=[]
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#上午第一節到第四節設1
                    self.ClassData[ClassIndx][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                    self.ClassData[ClassIndx][self.ClassData[0].index('課表')][0][LoopNumA]=7

    def TeacherConditionRange(self):
        for LoopIndex in range(1,len(self.TeacherData)):

            setTime=self.TeacherData[LoopIndex][self.TeacherData[0].index('上下午')]#判斷上下午是否可排課
            ClassSet=[[0 for col in range(6)]for row in range(8)]

            if setTime=='上午':
                for LoopNumA in range(1,6,1):#星期
                    for LoopNumB in range(1,5,1):#上午第一節到第四節設1
                        ClassSet[LoopNumB][LoopNumA]=1
            elif setTime=='下午':
                for LoopNumA in range(1,6,1):#星期
                    for LoopNumB in range(5,8,1):#下午第五節到第七節設1
                        ClassSet[LoopNumB][LoopNumA]=1
            else:
                for LoopNumA in range(1,6,1):#星期
                    for LoopNumB in range(1,8,1):#下午第五節到第七節設1
                        ClassSet[LoopNumB][LoopNumA]=1
            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#邏輯判定
                    if self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==1 and ClassSet[LoopNumB][LoopNumA]==1 :
                        self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:
                        self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=0


            setTime=self.TeacherData[LoopIndex][self.TeacherData[0].index('限制星期')]#判斷星期是否可排課
            setTime=str(setTime)
            ClassSet=[[0 for col in range(6)]for row in range(8)]

            if setTime!='12345':
                for LoopNumA in (setTime):
                    for LoopNumB in range(1,8,1):#星期整天設1
                        ClassSet[LoopNumB][int(LoopNumA)]=1
            else:
                for LoopNumA in range(1,6,1):#星期
                    for LoopNumB in range(1,8,1):#下午第五節到第七節設1
                        ClassSet[LoopNumB][LoopNumA]=1

            for LoopNumA in range(1,6,1):#星期
                for LoopNumB in range(1,8,1):#邏輯判定
                    if self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==1 and ClassSet[LoopNumB][LoopNumA]==1:
                        self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:
                        self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=0

            for SC in range(0,10,1):
                setTime=str(self.TeacherData[LoopIndex][self.LessonData[0].index('條件'+str(SC+1))])
                SCL=setTime.split('0')
                if SCL[0]!=0:
                    if SCL[0]=='可':
                        ClassSet=[[0 for col in range(6)]for row in range(8)]
                        for SCLDay in range(1,len(SCL),2):
                            for SCLTime in SCL[SCLDay+1]:
                                ClassSet[int(SCLTime)][int(SCL[int(SCLDay)])]=1

                    elif SCL[0]=='不':
                        ClassSet=[[1 for col in range(6)]for row in range(8)]
                        for SCLDay in range(1,len(SCL),2):
                            for SCLTime in SCL[SCLDay+1]:
                                ClassSet[int(SCLTime)][int(SCL[int(SCLDay)])]=0
                    for LoopNumA in range(1,6,1):#星期
                        for LoopNumB in range(1,8,1):#邏輯判定
                             if self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==0 or \
                             self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==1:
                                if self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                                ClassSet[LoopNumB][LoopNumA]==1:
                                    self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1
                                else:
                                    self.TeacherData[LoopIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=0

    def TeacherConditionSet(self,ClassIndex):
        for LoopNumA in range(1,6,1):#星期
            for LoopNumB in range(1,8,1):#邏輯判定
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0 or\
                 self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and \
                    self.TeacherData[self.TeacherIndex][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]==1 :
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=1
                    else:

                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]=0

    def getClass(self):#讀取Class資料

        return self.ClassData

    def getTeacherData(self):#讀取教師資料
        return self.TeacherData

    def PrintClassForm(self,ClassTemp):
        for SL in range(len(ClassTemp)):
            print(ClassTemp[SL])

    def testClassTemp(self,ClassIndex,LessonTime,LessonContinue):
        RGDayRecordIndex=[]
        RGDayRecordTimes=[]
        RGDayRecordIndex2=[]
        RGDayRecordTimes2=[]
        TempNO=0
        TF=False
        if LessonContinue==1:

            if LessonTime<=5:
                for LoopNumA in range(1,6,1):#測當天還剩多少空堂
                    CountET=0
                    for LoopNumB in range(1,8,1):
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                           CountET+=1
                    if CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                if  len(RGDayRecordIndex)>=LessonTime:
                    TF=True
            elif LessonTime>5:
                for LoopNumA in range(1,6,1):#測當天還剩多少空堂
                    CountET=0
                    CountET2=0
                    for LoopNumB in range(1,7,1):
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1\
                        and self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1:
                           CountET2+=1
                    if CountET2>0:
                        RGDayRecordIndex2.append(LoopNumA)
                        RGDayRecordTimes2.append(CountET)
                    for LoopNumB in range(1,8,1):
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                           CountET+=1
                    if CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                if  len(RGDayRecordIndex)==5 and len(RGDayRecordIndex2)>=LessonTime-5:
                    TF=True
            return TF
        elif LessonContinue==2:#連堂數2判別

            if LessonTime<=5:
                for LoopNumA in range(1,6,1):#測當天還剩多少空堂
                    CountET=0
                    for LoopNumB in range(1,7,1):
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and\
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1:
                           CountET+=1
                    if CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                if  len(RGDayRecordIndex)>=LessonTime/LessonContinue:
                    TF=True

            return TF
        elif LessonContinue==3:#連堂數3判別

            if LessonTime<=5:
                for LoopNumA in range(1,6,1):#測當天還剩多少空堂
                    CountET=0
                    for LoopNumB in (1,2,5):
                        if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1 and\
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+1][LoopNumA]==1 and\
                        self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB+2][LoopNumA]==1 :
                           CountET+=1
                    if CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                if  len(RGDayRecordIndex)>=LessonTime/LessonContinue:
                    TF=True
        elif LessonContinue==4:#連堂數3判別

            if LessonTime<=5:
                for LoopNumA in range(1,6,1):#測當天還剩多少空堂
                    CountET=0
                    if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1][LoopNumA]==1 and\
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+1][LoopNumA]==1 and\
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+2][LoopNumA]==1 and \
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][1+3][LoopNumA]==1 :
                       CountET+=1
                    if CountET>0:
                        RGDayRecordIndex.append(LoopNumA)
                        RGDayRecordTimes.append(CountET)
                if  len(RGDayRecordIndex)>=LessonTime/LessonContinue:
                    TF=True

            return TF
############################################################################################
    def ClearData(self,Data):
        for LoopNumA in range(len(Data)):
            if str(Data[LoopNumA][0])=='0' or str(Data[LoopNumA][0])=='None':
                del Data[LoopNumA:len(Data)]
                break
        return Data

    def CheckEmpty(self,ClassIndex):
        Empty=[]
        for LoopNumA in range(1,6,1):
            for LoopNumB in range(1,8,1):
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==1:
                    Empty.append([LoopNumB,LoopNumA])
        return Empty

    def CheckOccu(self,ClassIndex):
        Occu=[]
        for LoopNumA in range(1,6,1):
            for LoopNumB in range(1,8,1):
                if self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA]==0:
                    Occu.append([LoopNumB,LoopNumA])
        return Occu



    #def Switch(self,ClassIndex,TeacherIndex,Empty):
        #for LoopNumA in range(len(Empty)):
            #if self.TeacherData[TeacherIndex][self.TeacherData[0].index('課表')][Empty[LoopNumA][0]][Empty[LoopNumA][1]]==1:

    def SearchClass(self,ClassIndex,LessonYears,LessonTime,LessonName,LessonClass,Occu,ClassNO):
        ClassInfo=[]
        TeacherName=''
        self.ClassTemp=self.ClassData[ClassIndex][self.ClassData[0].index('課表')]
        LessonList=[]
        for LoopNumA in range(1,6,1):
            for LoopNumB in range(1,8,1):
               ClassInfo=self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopNumB][LoopNumA].split(',')
               TeacherName=ClassInfo[1]
               for LoopTeacher in range (len(self.TeacherData)):
                   if self.TeacherData[LoopTeacher][self.TeacherData[0].index('姓名')]==TeacherName and \
                    self.TeacherData[LoopTeacher][self.TeacherData[0].index('連堂')]==1 and \
                    self.TeacherData[LoopTeacher][self.TeacherData[0].index('節數')]<=5 :
                        if self.TeacherData[LoopTeacher][self.TeacherData[0].index('年級')]==LessonYears or \
                        self.TeacherData[LoopTeacher][self.TeacherData[0].index('年級')]==0:
                            if self.TeacherData[LoopTeacher][self.TeacherData[0].index('班級')]==LessonClass or \
                            self.TeacherData[LoopTeacher][self.TeacherData[0].index('班級')]==0:
                                if self.TeacherData[LoopTeacher][self.TeacherData[0].index('課表')][Occu[0]][Occu[1]]==1:
                                    self.TeacherData[LoopTeacher][self.TeacherData[0].index('課表')][LoopNumB][LoopNumA]=1
                                    self.TeacherData[LoopTeacher][self.TeacherData[0].index('課表')][Occu[0]][Occu[1]]=ClassNO
                                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][Occu[0]][Occu[1]]=ClassInfo

    def LessonFilter1(self,LessonInfo,LessonContinue,LessonYears):#只調連堂數1及5堂以下的課程
        SigleLessonInfo=[]
        NewLoopLessonInfo=[]
        for LoopLessonInfo in range(len(LessonInfo)):
            SigleLessonInfo=LessonInfo[LoopLessonInfo].split(',')
            for LoopLessonData in range(len(self.LessonData)):
                if self.LessonData[LoopLessonData][self.LessonData[0].index('科目')]==SigleLessonInfo[0] and\
                self.LessonData[LoopLessonData][self.LessonData[0].index('年級')]==LessonYears and\
                self.LessonData[LoopLessonData][self.LessonData[0].index('節數')]<=5 and\
                self.LessonData[LoopLessonData][self.LessonData[0].index('連堂數')]==LessonContinue:
                    NewLoopLessonInfo.append(LessonInfo[LoopLessonInfo])
        return NewLoopLessonInfo

    def LessonFilter2(self,LessonInfo,ClassIndex):
        Occu=self.CheckOccu(ClassIndex)
        OccuTemp=[]
        NewLoopLessonInfo=[]
        SigleLessonInfo=[]
        TF=True
        for LoopOccu in range(len(Occu)):#從不可排課的那節當天開始
            OccuTemp=str(Occu[LoopOccu]).split(',')
            for LoopLessonInfo in range(len(LessonInfo)):
                SigleLessonInfo=LessonInfo[LoopLessonInfo].split(',')
                for LoopTime in range(1,8,1):
                    self.ClassData[ClassIndex][self.ClassData[0].index('課表')][LoopTime][int(OccuTemp[1])]=1


    def SwitchTest(self,Occu,ClassIndex,TeacherIndex):
        ClassTemp=[]
        TeacherClassTemp1=[]
        TeacherClassTemp2=[]
        ClassTemp=self.ClassData[ClassIndex][self.ClassData[0].index('課表')].copy
        TeacherClassTemp1=self.TeacherData[TeacherIndex][TeacherData[0].index('課表')].copy()
if __name__ == '__main__':
    SeriaWord='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    Class=ExcelAction('.\\班級.xlsx')
    ClassSheetName,ClassSheet,ClassRow,ClassCol=Class.Loading(0)#載入班級資料
    ClassData=[[0 for col in range(ClassCol)] for row in range(ClassRow)]
    Lesson=ExcelAction('.\\課程.xlsx')
    LessonSheetName,LessonSheet,LessonRow,LessonCol=Lesson.Loading(0)#載入課程資料
    LessonData=[[0 for col in range(LessonCol)] for row in range(LessonRow)]
    Teacher=ExcelAction('.\\教師.xlsx')
    TeacherSheetName,TeacherSheet,TeacherRow,TeacherCol=Teacher.Loading(0)#載入教師資料
    TeacherData=[[0 for col in range(TeacherCol)] for row in range(TeacherRow)]
    TeacherDataOrigin=[[0 for col in range(TeacherCol)] for row in range(TeacherRow)]#原始資料存檔
    Student=ExcelAction('.\\學生.xlsx')


    SingleLesson=[]
    ClassTemp=ClassForm()
    TeacherTemp=ClassForm()

    for LoopNumA in range(0,LessonRow,1):
        for LoopNumB in range(LessonCol):
            LessonData[LoopNumA][LoopNumB]=LessonSheet[SeriaWord[LoopNumB]+str(LoopNumA+1)].value#讀取課程資料

    for LoopNumA in range(0,TeacherRow,1):
        for LoopNumB in range(TeacherCol):
            TeacherData[LoopNumA][LoopNumB]=TeacherSheet[SeriaWord[LoopNumB]+str(LoopNumA+1)].value#讀取教師資料
            TeacherDataOrigin[LoopNumA][LoopNumB]=TeacherSheet[SeriaWord[LoopNumB]+str(LoopNumA+1)].value
    for LoopNumA in range(0,ClassRow,1):
        for LoopNumB in range(ClassCol):
            ClassData[LoopNumA][LoopNumB]=ClassSheet[SeriaWord[LoopNumB]+str(LoopNumA+1)].value#讀取教師資料
    #TeacherDataOrigin=TeacherData.copy()
    Ranger=RangeTime(TeacherData,ClassTemp,ClassData,LessonData,TeacherDataOrigin)#開啟RangeTime Class

    LessonData=Ranger.ClearData(LessonData)#清除多餘的課程資料
    TeacherData=Ranger.ClearData(TeacherData)#清除多餘的教師資料
    ClassData=Ranger.ClearData(ClassData)#清除多餘的班級資料

    TeacherData=Ranger.reRangeTeacherData()#排列教師權數
    LessonData=Ranger.reRangeLesson()#排列課表權數
    for loop in range(len(LessonData)):
        print(LessonData[loop])

    Ranger.buildTeacherClassForm()#建構老師課表
    Ranger.buildClassClassForm()#建構班級課表

    Ranger.TeacherConditionRange()#設定老師條件課表
    TeacherData=Ranger.getTeacherData()#取出老師Data
    ClassNO=''
    for LoopClass in range(1,len(ClassData)):
        Ranger.reNewClass(LoopClass)
    TF=False#判別True與False用

    TFCount=0
    LessonCount=0
    #for test in range(100):
    while LessonCount<len(LessonData)-1:
        TFCount=1
        LessonCount=0
        for LoopLesson in range(1,len(LessonData)):
            LessonCount+=1
            LessonYears=LessonData[LoopLesson][LessonData[0].index('年級')]#獲取科目年級
            LessonName=LessonData[LoopLesson][LessonData[0].index('科目')]#獲取科目名
            LessonTime=LessonData[LoopLesson][LessonData[0].index('節數')]#獲取節數
            LessonClass=LessonData[LoopLesson][LessonData[0].index('班級')]#獲取班級
            LessonContinue=LessonData[LoopLesson][LessonData[0].index('連堂數')]#獲取連堂數
            TF=False
            while TF==False:#測試教師是否符合
                for LoopClass in range(1,len(ClassData)):
                    Ranger.reSetClass(LoopClass)#將課表更新
                    ClassYears=ClassData[LoopClass][ClassData[0].index('年級')]
                    ClassClass=ClassData[LoopClass][ClassData[0].index('班級')]
                    if ClassYears==LessonYears:
                        ClassNO=str(ClassYears)+'年'+str(ClassClass)+'班'
                        print(ClassNO,LessonName)
                        Ranger.reSetClass(LoopClass)
                        Ranger.LessonConditionSet(LoopLesson,LoopClass)
                        TeacherIndex=Ranger.getTeacher(LessonName,LessonTime,LessonYears,LessonClass)
                        Ranger.TeacherConditionSet(LoopClass)
                        TF=Ranger.testClassTemp(LoopClass,LessonTime,LessonContinue)
                        if TF==False:
                            #print('Bug Check 課表比較')
                            testNum=1
                            if testNum==1:
                                Ranger.PrintClassForm(ClassData[LoopClass][ClassData[0].index('課表')])
                                print(TeacherData[TeacherIndex][TeacherData[0].index('姓名')])
                                Ranger.PrintClassForm(TeacherData[TeacherIndex][TeacherData[0].index('課表')])
                                print(ClassData[LoopClass][ClassData[0].index('科目教師')])
                                os.system('pause')

                            Ranger.reLessonAllTeacher(LessonName,LessonYears,ClassYears)
                            Ranger.reLessonAllClass(LessonName,LessonYears)
                            TFCount+=1
                            LessonCount-=1
                            break
                        if LessonContinue==1:
                            Ranger.RangeLesson1(LessonName,LessonTime,ClassNO,LoopClass)
                        elif LessonContinue==2:
                            Ranger.RangeLesson2(LessonName,LessonTime,ClassNO,LoopClass)
                        elif LessonContinue==3:
                            Ranger.RangeLesson3(LessonName,LessonTime,ClassNO,LoopClass)
                        elif LessonContinue==4:
                            Ranger.RangeLesson4(LessonName,LessonTime,ClassNO,LoopClass)
                        ClassData=Ranger.getClass()
                        #print(ClassData[LoopClass][ClassData[0].index('科目教師')])


                if TFCount>200:
                    break
            if TFCount>200:
                Ranger.reNewLessonAllClass()
                Ranger.reNewLessonAllTeacher()

                break
    ClassData=Ranger.getClass()
    TeacherData=Ranger.getTeacherData()
    for LoopLesson in range(1,len(LessonData)):
        Ranger.PrintClassForm(ClassData[LoopClass][ClassData[0].index('課表')])








